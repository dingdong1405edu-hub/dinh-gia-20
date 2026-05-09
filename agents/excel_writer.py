"""Agent 8b: Xuất dữ liệu định giá ra file Excel (openpyxl).

Sheets:
  1. Tổng quan         - Company / period / fair value summary
  2. KQKD              - Income Statement (current + previous)
  3. BCĐKT             - Balance Sheet (current + previous)
  4. LCTT              - Cash flow (nếu BCTC có)
  5. Tỷ số             - 17 tỷ số (current + previous + tăng trưởng YoY)
  6. Dự phóng 5Y       - Doanh thu / EBITDA / FCFF 5 năm + giả định
  7. Định giá          - DCF + Multiples + fair value
  8. Sensitivity       - Sensitivity matrix WACC × Terminal growth
  9. Phân tích ngành   - TAM/SAM/SOM, đối thủ, drivers, risks
  10. Investment Thesis - Headline / drivers / catalysts / risks / deal

Schemas tương ứng:
  - extractor.py: balance_sheet.{current,previous}.{assets|liabilities|equity}
                  income_statement.{current,previous}.{revenue,...,net_profit_after_tax}
                  cash_flow.{current,previous}.{cf_operating,cf_investing,...}
                  period.{current,previous}.label
  - projector.py:  projection.projections[i].{revenue, growth_pct, ebit, ebitda,
                   net_income, capex, change_in_wc, fcff, ebitda_margin_pct, ...}
  - valuator.py:   valuation.{assumptions, dcf, multiples, sensitivity, summary}
                   sensitivity.matrix = [{wacc_pct, values: [{terminal_growth_pct, equity_value}]}]
                   summary.method_values = [{method, equity_value}]
                   multiples.<key> = {multiple, equity_value, ...}  (NO median/min/max)
  - analyzer.py:   ratios = {current, previous, changes}; growth = {revenue_yoy, ...}
"""
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agents.report_style import COLORS

# ---------- Styles (đồng bộ với report_style) ----------
def _hex(c: str) -> str:
    return c.lstrip("#").upper()


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=_hex(COLORS["primary"]))
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=_hex(COLORS["primary"]))
SECTION_FILL = PatternFill("solid", fgColor=_hex(COLORS["primary_light"]))
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=_hex(COLORS["text_strong"]))
BODY_FONT = Font(name="Calibri", size=10, color=_hex(COLORS["text"]))
MUTED_FONT = Font(name="Calibri", size=9, italic=True, color=_hex(COLORS["text_muted"]))
TOTAL_FILL = PatternFill("solid", fgColor=_hex(COLORS["surface_alt"]))
GOOD_FILL = PatternFill("solid", fgColor=_hex(COLORS["good_light"]))
WARN_FILL = PatternFill("solid", fgColor=_hex(COLORS["warning_light"]))
POOR_FILL = PatternFill("solid", fgColor=_hex(COLORS["poor_light"]))
BASE_HIGHLIGHT_FILL = PatternFill("solid", fgColor=_hex(COLORS["primary_light"]))

THIN_SIDE = Side(style="thin", color=_hex(COLORS["border"]))

INPUT_FILL = PatternFill("solid", fgColor=_hex(COLORS["warning_light"]))  # màu vàng cho ô input edit-được
INPUT_FONT = Font(name="Calibri", size=11, bold=True, color=_hex(COLORS["text_strong"]))
FORMULA_TAG_FONT = Font(name="Consolas", size=9, italic=True, color=_hex(COLORS["text_muted"]))


# ---------- Cross-sheet cell ref tracker ----------
class CellRefs:
    """Ghi nhận vị trí ô của các giá trị nguồn (BCĐKT, KQKD, Dự phóng…) để các
    sheet tính toán (Tỷ số, Định giá) viết FORMULA tham chiếu thật, không phải
    số tĩnh.

    Mục đích: kế toán bấm vào ô tỷ số → thanh formula bar hiện
        ='BCĐKT'!B7 / 'BCĐKT'!B17
    để audit + sửa tay được.
    """
    def __init__(self):
        self._refs: dict[str, str] = {}  # name → "'Sheet Name'!B5"

    def record(self, name: str, sheet_title: str, coord: str) -> None:
        # Wrap sheet title in single quotes — required when title has spaces / VN chars.
        self._refs[name] = f"'{sheet_title}'!{coord}"

    def get(self, name: str) -> str | None:
        return self._refs.get(name)

    def has(self, *names: str) -> bool:
        """True iff TẤT CẢ tên đều đã được ghi nhận."""
        return all(n in self._refs for n in names)


def _coord(col: int, row: int) -> str:
    """1-based (col, row) → A1 string."""
    return f"{get_column_letter(col)}{row}"


# ---------- Public API ----------
def export_excel(payload: dict, output_path: str) -> dict:
    """Build XLSX from full pipeline payload. Returns metadata dict."""
    t0 = time.time()
    financials = (payload.get("extracted") or {}).get("financials") or {}
    industry = (payload.get("industry") or {}).get("industry") or {}
    business = (payload.get("business") or {}).get("business") or {}
    ratios_payload = payload.get("ratios") or {}
    projection = (payload.get("projection") or {}).get("projection") or {}
    valuation = payload.get("valuation") or {}
    thesis = (payload.get("thesis") or {}).get("thesis") or {}

    wb = Workbook()
    wb.remove(wb.active)

    # Cell-ref tracker để các sheet sau viết FORMULA tham chiếu các ô đã ghi.
    refs = CellRefs()

    # Build theo thứ tự dependency: source sheets (BCĐKT, KQKD, Dự phóng) phải
    # viết TRƯỚC computed sheets (Tỷ số, Định giá). Tab order trong UI đúng theo
    # creation order.
    _safe(_sheet_methodology, wb)
    _safe(_sheet_overview, wb, financials, business, valuation, thesis)
    _safe(_sheet_income_statement, wb, financials, refs)
    _safe(_sheet_balance_sheet, wb, financials, refs)
    _safe(_sheet_cash_flow, wb, financials)
    _safe(_sheet_ratios, wb, ratios_payload, refs)
    _safe(_sheet_projection, wb, projection, financials.get("unit"), refs)
    _safe(_sheet_valuation, wb, valuation, financials.get("unit"), refs)
    _safe(_sheet_sensitivity, wb, valuation)
    _safe(_sheet_industry, wb, industry)
    _safe(_sheet_thesis, wb, thesis)

    # Đảm bảo workbook luôn có ít nhất 1 sheet (openpyxl yêu cầu).
    if not wb.worksheets:
        ws = wb.create_sheet("Empty")
        ws.cell(1, 1, "Không có dữ liệu khả dụng để xuất.").font = MUTED_FONT

    wb.save(output_path)
    return {
        "elapsed_sec": round(time.time() - t0, 2),
        "output_path": output_path,
        "size_bytes": Path(output_path).stat().st_size,
        "sheets": [s.title for s in wb.worksheets],
    }


def _safe(builder, wb, *args, **kwargs):
    """Wrap a sheet builder so a schema-drift error in one sheet doesn't kill the export."""
    try:
        builder(wb, *args, **kwargs)
    except Exception as exc:
        title = builder.__name__.replace("_sheet_", "").replace("_", " ").title()
        ws = wb.create_sheet(f"⚠ {title[:25]}")
        ws.cell(1, 1, f"Sheet '{title}' lỗi khi xuất:").font = LABEL_FONT
        ws.cell(2, 1, repr(exc)).font = MUTED_FONT
        ws.column_dimensions["A"].width = 90


# ---------- Sheet builders ----------
def _sheet_overview(wb, financials, business, valuation, thesis):
    ws = wb.create_sheet("Tổng quan")
    company = financials.get("company") or {}
    period = financials.get("period") or {}
    cur_label = (period.get("current") or {}).get("label") or "—"
    prev_label = (period.get("previous") or {}).get("label") or "—"
    summary = valuation.get("summary") or {}
    es = thesis.get("executive_summary") or {}

    rows = [
        ("Tên doanh nghiệp", company.get("name")),
        ("Mã số thuế", company.get("tax_code")),
        ("Địa chỉ", company.get("address")),
        ("Ngành (BCTC)", company.get("industry")),
        ("Loại báo cáo", company.get("report_type")),
        ("Đơn vị", financials.get("unit") or "đồng"),
        ("Kỳ hiện tại", cur_label),
        ("Kỳ trước", prev_label),
        ("Vị thế cạnh tranh", business.get("competitive_position")),
        ("Giai đoạn tăng trưởng", business.get("growth_stage")),
    ]
    _write_section_title(ws, 1, 1, "THÔNG TIN DOANH NGHIỆP", span=2)
    _write_kv_block(ws, rows, start_row=3)

    next_row = 3 + len(rows) + 1
    _write_section_title(ws, next_row, 1, "TÓM TẮT ĐỊNH GIÁ", span=2)
    val_rows = [
        ("Fair value low", summary.get("fair_value_low")),
        ("Fair value mid", summary.get("fair_value_mid")),
        ("Fair value high", summary.get("fair_value_high")),
        (f"Sau chiết khấu thiểu số ({summary.get('minority_discount_pct') or 0}%)",
         summary.get("fair_value_after_minority_discount")),
    ]
    _write_kv_block(ws, val_rows, start_row=next_row + 2, value_format="#,##0")

    next_row = next_row + 2 + len(val_rows) + 1
    if es.get("headline") or es.get("recommendation"):
        _write_section_title(ws, next_row, 1, "EXECUTIVE SUMMARY", span=2)
        ws.cell(next_row + 2, 1, "Headline").font = LABEL_FONT
        c = ws.cell(next_row + 2, 2, es.get("headline") or "—")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = BODY_FONT
        ws.cell(next_row + 3, 1, "Khuyến nghị").font = LABEL_FONT
        c = ws.cell(next_row + 3, 2, es.get("recommendation") or "—")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = BODY_FONT

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A2"


def _sheet_income_statement(wb, financials, refs=None):
    ws = wb.create_sheet("KQKD")
    is_data = financials.get("income_statement") or {}
    cur = is_data.get("current") or {}
    prev = is_data.get("previous") if isinstance(is_data.get("previous"), dict) else {}
    period = financials.get("period") or {}
    cur_label = (period.get("current") or {}).get("label") or "Kỳ này"
    prev_label = (period.get("previous") or {}).get("label") or "Kỳ trước"

    # (key, label, ref_name) — ref_name=None nếu không cần track cho công thức.
    fields = [
        ("revenue", "Doanh thu bán hàng và cung cấp dịch vụ", None),
        ("revenue_deductions", "Các khoản giảm trừ doanh thu", None),
        ("net_revenue", "Doanh thu thuần", "net_revenue"),
        ("cogs", "Giá vốn hàng bán", "cogs"),
        ("gross_profit", "Lợi nhuận gộp", "gross_profit"),
        ("financial_income", "Doanh thu hoạt động tài chính", None),
        ("financial_expense", "Chi phí tài chính", None),
        ("interest_expense", "  Trong đó: Chi phí lãi vay", "interest_expense"),
        ("selling_expense", "Chi phí bán hàng", None),
        ("admin_expense", "Chi phí quản lý doanh nghiệp", None),
        ("operating_profit", "Lợi nhuận thuần từ HĐKD (EBIT)", "operating_profit"),
        ("other_income", "Thu nhập khác", None),
        ("other_expense", "Chi phí khác", None),
        ("profit_before_tax", "Tổng lợi nhuận kế toán trước thuế", "profit_before_tax"),
        ("current_tax", "Chi phí thuế TNDN hiện hành", None),
        ("deferred_tax", "Chi phí thuế TNDN hoãn lại", None),
        ("net_profit_after_tax", "Lợi nhuận sau thuế TNDN", "net_profit_after_tax"),
        ("eps", "EPS", None),
    ]
    _write_financial_table(ws, fields, cur, prev, cur_label, prev_label, refs=refs, sheet_title=ws.title)
    ws.freeze_panes = "B2"


def _sheet_balance_sheet(wb, financials, refs=None):
    ws = wb.create_sheet("BCĐKT")
    bs = financials.get("balance_sheet") or {}
    cur = bs.get("current") or {}
    prev = bs.get("previous") if isinstance(bs.get("previous"), dict) else {}
    period = financials.get("period") or {}
    cur_label = (period.get("current") or {}).get("label") or "Kỳ này"
    prev_label = (period.get("previous") or {}).get("label") or "Kỳ trước"

    a_cur = cur.get("assets") or {}
    a_prev = prev.get("assets") or {} if isinstance(prev, dict) else {}
    l_cur = cur.get("liabilities") or {}
    l_prev = prev.get("liabilities") or {} if isinstance(prev, dict) else {}
    e_cur = cur.get("equity") or {}
    e_prev = prev.get("equity") or {} if isinstance(prev, dict) else {}

    # Mỗi leaf row có thêm trường thứ 5 = ref_name (None nếu không track).
    sections = [
        ("TÀI SẢN", [
            ("Tài sản ngắn hạn", "subheader"),
            ("Tiền và TĐ tiền",       a_cur.get("cash_and_equivalents"),       a_prev.get("cash_and_equivalents"),       "leaf",  "cash_and_equivalents"),
            ("Đầu tư TC ngắn hạn",    a_cur.get("short_term_investments"),     a_prev.get("short_term_investments"),     "leaf",  "short_term_investments"),
            ("Phải thu ngắn hạn",     a_cur.get("short_term_receivables"),     a_prev.get("short_term_receivables"),     "leaf",  "short_term_receivables"),
            ("Hàng tồn kho",          a_cur.get("inventory"),                  a_prev.get("inventory"),                  "leaf",  "inventory"),
            ("TS ngắn hạn khác",      a_cur.get("other_current_assets"),       a_prev.get("other_current_assets"),       "leaf",  None),
            ("Tổng tài sản ngắn hạn", a_cur.get("current_assets_total"),       a_prev.get("current_assets_total"),       "total", "current_assets_total"),
            ("Tài sản dài hạn", "subheader"),
            ("Phải thu dài hạn",      a_cur.get("long_term_receivables"),      a_prev.get("long_term_receivables"),      "leaf",  None),
            ("Tài sản cố định",       a_cur.get("fixed_assets"),               a_prev.get("fixed_assets"),               "leaf",  None),
            ("Bất động sản đầu tư",   a_cur.get("investment_properties"),      a_prev.get("investment_properties"),      "leaf",  None),
            ("Đầu tư TC dài hạn",     a_cur.get("long_term_investments"),      a_prev.get("long_term_investments"),      "leaf",  None),
            ("TS dài hạn khác",       a_cur.get("other_non_current_assets"),   a_prev.get("other_non_current_assets"),   "leaf",  None),
            ("Tổng tài sản dài hạn",  a_cur.get("non_current_assets_total"),   a_prev.get("non_current_assets_total"),   "total", "non_current_assets_total"),
            ("TỔNG CỘNG TÀI SẢN",     a_cur.get("total_assets"),               a_prev.get("total_assets"),               "grand", "total_assets"),
        ]),
        ("NGUỒN VỐN", [
            ("Nợ phải trả ngắn hạn", "subheader"),
            ("Vay & nợ ngắn hạn",     l_cur.get("short_term_debt"),                l_prev.get("short_term_debt"),                "leaf",  "short_term_debt"),
            ("Phải trả người bán",    l_cur.get("accounts_payable"),               l_prev.get("accounts_payable"),               "leaf",  None),
            ("Nợ ngắn hạn khác",      l_cur.get("other_current_liabilities"),      l_prev.get("other_current_liabilities"),      "leaf",  None),
            ("Tổng nợ ngắn hạn",      l_cur.get("current_liabilities_total"),      l_prev.get("current_liabilities_total"),      "total", "current_liabilities_total"),
            ("Nợ phải trả dài hạn", "subheader"),
            ("Vay & nợ dài hạn",      l_cur.get("long_term_debt"),                 l_prev.get("long_term_debt"),                 "leaf",  "long_term_debt"),
            ("Nợ dài hạn khác",       l_cur.get("other_non_current_liabilities"),  l_prev.get("other_non_current_liabilities"),  "leaf",  None),
            ("Tổng nợ dài hạn",       l_cur.get("non_current_liabilities_total"),  l_prev.get("non_current_liabilities_total"),  "total", "non_current_liabilities_total"),
            ("TỔNG NỢ PHẢI TRẢ",      l_cur.get("total_liabilities"),              l_prev.get("total_liabilities"),              "grand", "total_liabilities"),
            ("Vốn chủ sở hữu", "subheader"),
            ("Vốn góp CSH",                e_cur.get("share_capital"),     e_prev.get("share_capital"),     "leaf",  None),
            ("LN sau thuế chưa phân phối", e_cur.get("retained_earnings"), e_prev.get("retained_earnings"), "leaf",  None),
            ("VCSH khác",                  e_cur.get("other_equity"),      e_prev.get("other_equity"),      "leaf",  None),
            ("TỔNG VỐN CHỦ SỞ HỮU",        e_cur.get("total_equity"),      e_prev.get("total_equity"),      "grand", "total_equity"),
        ]),
    ]

    headers = ["Khoản mục", cur_label, prev_label, "Δ tuyệt đối", "Δ %"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center" if col > 1 else "left")
    row = 2
    for section_title, items in sections:
        c = ws.cell(row, 1, section_title)
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
        for col in range(2, 6):
            ws.cell(row, col).fill = SECTION_FILL
        row += 1
        for item in items:
            label = item[0]
            if len(item) == 2 and item[1] == "subheader":
                cell = ws.cell(row, 1, label)
                cell.font = Font(name="Calibri", size=10, bold=True, italic=True,
                                 color=_hex(COLORS["text_muted"]))
                row += 1
                continue
            cv = item[1]; pv = item[2]
            level = item[3] if len(item) > 3 else "leaf"
            ref_name = item[4] if len(item) > 4 else None
            label_font = LABEL_FONT if level in ("total", "grand") else BODY_FONT
            ws.cell(row, 1, label).font = label_font

            cv_cell = None; pv_cell = None
            if isinstance(cv, (int, float)):
                cv_cell = ws.cell(row, 2, float(cv))
                cv_cell.number_format = "#,##0"
                cv_cell.font = label_font
            if isinstance(pv, (int, float)):
                pv_cell = ws.cell(row, 3, float(pv))
                pv_cell.number_format = "#,##0"
                pv_cell.font = label_font

            # Δ là FORMULA — bấm vào ô sẽ thấy =B-C
            if cv_cell is not None and pv_cell is not None:
                d_abs = ws.cell(row, 4, f"={cv_cell.coordinate}-{pv_cell.coordinate}")
                d_abs.number_format = "#,##0"
                d_abs.font = label_font
                d_pct = ws.cell(row, 5,
                    f"=IFERROR(({cv_cell.coordinate}-{pv_cell.coordinate})"
                    f"/ABS({pv_cell.coordinate}),\"\")")
                d_pct.number_format = "+0.0%;-0.0%;0.0%"
                d_pct.font = label_font

            if level == "grand":
                for col in range(1, 6):
                    cell = ws.cell(row, col)
                    cell.fill = TOTAL_FILL
                    cell.font = LABEL_FONT

            # Track ref cho computed sheets tham chiếu.
            if ref_name and refs is not None and cv_cell is not None:
                refs.record(ref_name, ws.title, cv_cell.coordinate)

            row += 1
        row += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "B2"


def _sheet_cash_flow(wb, financials):
    cf = financials.get("cash_flow") or {}
    cur = cf.get("current") or {}
    prev = cf.get("previous") if isinstance(cf.get("previous"), dict) else {}
    if not cur and not prev:
        return  # bỏ sheet nếu không có LCTT trong BCTC

    ws = wb.create_sheet("LCTT")
    period = financials.get("period") or {}
    cur_label = (period.get("current") or {}).get("label") or "Kỳ này"
    prev_label = (period.get("previous") or {}).get("label") or "Kỳ trước"
    fields = [
        ("cf_operating", "Lưu chuyển tiền thuần từ HĐKD", None),
        ("cf_investing", "Lưu chuyển tiền thuần từ HĐ đầu tư", None),
        ("cf_financing", "Lưu chuyển tiền thuần từ HĐ tài chính", None),
        ("net_cf", "Lưu chuyển tiền thuần trong kỳ", None),
        ("ending_cash", "Tiền và tương đương tiền cuối kỳ", None),
    ]
    _write_financial_table(ws, fields, cur, prev, cur_label, prev_label)
    ws.freeze_panes = "B2"


def _sheet_ratios(wb, ratios_payload, refs=None):
    """Tỷ số tài chính — LIVE FORMULAS tham chiếu BCĐKT/KQKD.

    Khi kế toán bấm vào ô tỷ số sẽ thấy formula bar hiện vd:
        ='BCĐKT'!B7 / 'BCĐKT'!B17
    Ô tỷ số cập nhật tự động nếu sửa BCĐKT/KQKD.

    Khi extractor không có đủ trường nguồn → fallback ghi giá trị tĩnh từ analyzer.
    """
    ws = wb.create_sheet("Tỷ số")
    ratios = ratios_payload.get("ratios") or {}
    cur_flat = _flatten_ratios(ratios.get("current"))
    prev_flat = _flatten_ratios(ratios.get("previous") or ratios.get("prior"))
    growth = ratios_payload.get("growth") or {}

    def R(name: str) -> str | None:
        return refs.get(name) if refs else None

    # (key, label_VI, kind, refs_needed, formula_template, formula_text_human).
    # formula_template uses Python str.format with refs as kwargs.
    # kind: "pct_frac" = fraction display as %, "ratio" = plain number.
    PY = "_"  # placeholder when fallback only.
    categories = [
        ("Thanh khoản", [
            ("current_ratio", "Hệ số TT hiện hành", "ratio",
             ["current_assets_total", "current_liabilities_total"],
             "={current_assets_total} / {current_liabilities_total}",
             "= current_assets_total / current_liabilities_total"),
            ("quick_ratio", "Hệ số TT nhanh", "ratio",
             ["current_assets_total", "inventory", "current_liabilities_total"],
             "=({current_assets_total} - {inventory}) / {current_liabilities_total}",
             "= (current_assets_total − inventory) / current_liabilities_total"),
            ("cash_ratio", "Hệ số TT tiền mặt", "ratio",
             ["cash_and_equivalents", "short_term_investments", "current_liabilities_total"],
             "=({cash_and_equivalents} + {short_term_investments}) / {current_liabilities_total}",
             "= (cash + short_term_investments) / current_liabilities_total"),
        ]),
        ("Đòn bẩy / Cơ cấu vốn", [
            ("debt_ratio", "Hệ số nợ / TS", "pct_frac",
             ["total_liabilities", "total_assets"],
             "={total_liabilities} / {total_assets}",
             "= total_liabilities / total_assets"),
            ("debt_to_equity", "Nợ / VCSH", "ratio",
             ["total_liabilities", "total_equity"],
             "={total_liabilities} / {total_equity}",
             "= total_liabilities / total_equity"),
            ("equity_multiplier", "Hệ số nhân VCSH", "ratio",
             ["total_assets", "total_equity"],
             "={total_assets} / {total_equity}",
             "= total_assets / total_equity"),
            ("interest_coverage", "Khả năng trả lãi vay", "ratio",
             ["operating_profit", "interest_expense"],
             "={operating_profit} / {interest_expense}",
             "= operating_profit / interest_expense"),
            ("debt_to_ebitda", "Nợ / EBITDA", "ratio",
             ["total_liabilities", "operating_profit"],
             "={total_liabilities} / {operating_profit}",
             "= total_liabilities / EBITDA*  (* EBIT proxy nếu thiếu D&A)"),
        ]),
        ("Khả năng sinh lời", [
            ("gross_margin", "Biên LN gộp", "pct_frac",
             ["gross_profit", "net_revenue"],
             "={gross_profit} / {net_revenue}",
             "= gross_profit / net_revenue"),
            ("operating_margin", "Biên LN HĐKD", "pct_frac",
             ["operating_profit", "net_revenue"],
             "={operating_profit} / {net_revenue}",
             "= operating_profit / net_revenue"),
            ("ebitda_margin", "Biên EBITDA", "pct_frac",
             ["operating_profit", "net_revenue"],
             "={operating_profit} / {net_revenue}",
             "= EBITDA* / net_revenue  (* EBIT proxy nếu thiếu D&A)"),
            ("net_margin", "Biên LN ròng", "pct_frac",
             ["net_profit_after_tax", "net_revenue"],
             "={net_profit_after_tax} / {net_revenue}",
             "= net_profit_after_tax / net_revenue"),
            ("roa", "ROA", "pct_frac",
             ["net_profit_after_tax", "total_assets"],
             "={net_profit_after_tax} / {total_assets}",
             "= net_profit_after_tax / total_assets"),
            ("roe", "ROE", "pct_frac",
             ["net_profit_after_tax", "total_equity"],
             "={net_profit_after_tax} / {total_equity}",
             "= net_profit_after_tax / total_equity"),
        ]),
        ("Hiệu quả hoạt động", [
            ("asset_turnover", "Vòng quay tổng TS", "ratio",
             ["net_revenue", "total_assets"],
             "={net_revenue} / {total_assets}",
             "= net_revenue / total_assets"),
            ("inventory_turnover", "Vòng quay HTK", "ratio",
             ["cogs", "inventory"],
             "={cogs} / {inventory}",
             "= cogs / inventory"),
            ("receivables_turnover", "Vòng quay phải thu", "ratio",
             ["net_revenue", "short_term_receivables"],
             "={net_revenue} / {short_term_receivables}",
             "= net_revenue / short_term_receivables"),
        ]),
    ]

    headers = ["Tỷ số", "Kỳ này", "Kỳ trước", "Δ", "Công thức"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center" if col > 1 and col < 5 else "left")

    formula_text_font = Font(name="Consolas", size=9, color=_hex(COLORS["primary"]))
    row = 2
    cur_cells: dict[str, str] = {}  # ratio_name -> coordinate, để cột Δ tham chiếu kỳ trước nếu có.
    for cat_name, items in categories:
        c = ws.cell(row, 1, cat_name)
        c.font = SECTION_FONT; c.fill = SECTION_FILL
        for col in range(2, 6):
            ws.cell(row, col).fill = SECTION_FILL
        row += 1
        for key, label, kind, refs_needed, formula_tmpl, formula_text in items:
            cv = cur_flat.get(key)
            pv = prev_flat.get(key)
            ws.cell(row, 1, label).font = BODY_FONT
            fmt = "0.00%" if kind == "pct_frac" else "0.00"

            cv_cell = ws.cell(row, 2)
            # ⭐ Cột "Kỳ này" — ưu tiên FORMULA tham chiếu BCĐKT/KQKD; fallback giá trị tĩnh.
            if refs and refs.has(*refs_needed):
                kwargs = {n: refs.get(n) for n in refs_needed}
                cv_cell.value = formula_tmpl.format(**kwargs)
                cv_cell.number_format = fmt
                cv_cell.font = BODY_FONT
                cur_cells[key] = cv_cell.coordinate
            elif isinstance(cv, (int, float)):
                cv_cell.value = float(cv)
                cv_cell.number_format = fmt
                cv_cell.font = BODY_FONT
                cur_cells[key] = cv_cell.coordinate

            # Cột "Kỳ trước" — chỉ giá trị tĩnh (analyzer đã tính từ BCĐKT previous).
            if isinstance(pv, (int, float)):
                pv_cell = ws.cell(row, 3, float(pv))
                pv_cell.number_format = fmt
                pv_cell.font = BODY_FONT
                # Δ = formula tham chiếu trực tiếp 2 ô cùng sheet.
                if cur_cells.get(key):
                    d_cell = ws.cell(row, 4,
                        f"=IFERROR(({cur_cells[key]}-{pv_cell.coordinate})"
                        f"/ABS({pv_cell.coordinate}),\"\")")
                    d_cell.number_format = "+0.0%;-0.0%;0.0%"
                    d_cell.font = BODY_FONT

            f_cell = ws.cell(row, 5, formula_text)
            f_cell.font = formula_text_font
            f_cell.alignment = Alignment(wrap_text=True, vertical="center")
            row += 1
        row += 1

    # Growth block (giá trị tĩnh — không có cách viết formula tham chiếu vì BCĐKT
    # chỉ có 1 cột current; previous là cột riêng).
    if growth:
        c = ws.cell(row, 1, "Tăng trưởng YoY")
        c.font = SECTION_FONT; c.fill = SECTION_FILL
        for col in range(2, 6):
            ws.cell(row, col).fill = SECTION_FILL
        row += 1
        growth_items = [
            ("revenue_yoy", "Tăng trưởng doanh thu", "net_revenue", "KQKD",
             "= ('KQKD'!cur − 'KQKD'!prev) / ABS('KQKD'!prev)"),
            ("gross_profit_yoy", "Tăng trưởng LN gộp", "gross_profit", "KQKD",
             "= ('KQKD'!cur − 'KQKD'!prev) / ABS('KQKD'!prev)"),
            ("ebit_yoy", "Tăng trưởng EBIT", "operating_profit", "KQKD",
             "= ('KQKD'!cur − 'KQKD'!prev) / ABS('KQKD'!prev)"),
            ("net_income_yoy", "Tăng trưởng LNST", "net_profit_after_tax", "KQKD",
             "= ('KQKD'!cur − 'KQKD'!prev) / ABS('KQKD'!prev)"),
            ("total_assets_yoy", "Tăng trưởng tổng tài sản", "total_assets", "BCĐKT",
             "= ('BCĐKT'!cur − 'BCĐKT'!prev) / ABS('BCĐKT'!prev)"),
            ("total_equity_yoy", "Tăng trưởng VCSH", "total_equity", "BCĐKT",
             "= ('BCĐKT'!cur − 'BCĐKT'!prev) / ABS('BCĐKT'!prev)"),
        ]
        for k, label, _ref_name, _src, formula_text in growth_items:
            v = growth.get(k)
            ws.cell(row, 1, label).font = BODY_FONT
            if isinstance(v, (int, float)):
                cc = ws.cell(row, 2, float(v))
                cc.number_format = "+0.0%;-0.0%;0.0%"
                cc.font = BODY_FONT
            f_cell = ws.cell(row, 5, formula_text)
            f_cell.font = formula_text_font
            f_cell.alignment = Alignment(wrap_text=True, vertical="center")
            row += 1
        row += 1

    # Footer pointer + giải thích live formula.
    ws.cell(row + 1, 1, "💡 Cột 'Kỳ này' là LIVE FORMULA: bấm vào ô để thấy =B5/B17 — "
                       "sửa BCĐKT/KQKD thì tỷ số tự update. Cột 'Công thức' là chú giải text.").font = MUTED_FONT
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=5)
    ws.cell(row + 2, 1, "📁 Sửa công thức Python tại: agents/analyzer.py · _compute_period_ratios (line 84) · _compute_growth (line 200)").font = MUTED_FONT
    ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=5)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50
    ws.freeze_panes = "B2"


def _sheet_projection(wb, projection, unit, refs=None):
    ws = wb.create_sheet("Dự phóng 5Y")
    projections = projection.get("projections") or []
    if not projections:
        ws.cell(1, 1, "Không có dữ liệu dự phóng.").font = MUTED_FONT
        ws.column_dimensions["A"].width = 60
        return

    headers = ["Khoản mục"] + [
        p.get("year_label") or f"Y{p.get('year_index') or i+1}"
        for i, p in enumerate(projections)
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center" if col > 1 else "left")

    # (key, label, format, track_per_year). track_per_year=True → ghi nhận
    # ô từng năm cho DCF/Multiples tham chiếu công thức.
    fields = [
        ("revenue",            "Doanh thu",           "money",       False),
        ("growth_pct",         "Tăng trưởng DT (%)",  "pct_already", False),
        ("cogs",               "Giá vốn",             "money",       False),
        ("gross_profit",       "Lợi nhuận gộp",       "money",       False),
        ("operating_expense",  "OPEX",                "money",       False),
        ("ebit",               "EBIT",                "money",       False),
        ("depreciation",       "Khấu hao (D&A)",      "money",       False),
        ("ebitda",             "EBITDA",              "money",       True),  # cho EV/EBITDA multiple
        ("ebitda_margin_pct",  "Biên EBITDA (%)",     "pct_already", False),
        ("interest_expense",   "Chi phí lãi vay",     "money",       False),
        ("profit_before_tax",  "LN trước thuế",       "money",       False),
        ("tax",                "Thuế TNDN",           "money",       False),
        ("net_income",         "LNST",                "money",       False),
        ("net_margin_pct",     "Biên LN ròng (%)",    "pct_already", False),
        ("capex",              "Capex",               "money",       False),
        ("change_in_wc",       "Δ Vốn lưu động",      "money",       False),
        ("fcff",               "FCFF (Free CF firm)", "money",       True),  # cho DCF chiết khấu
    ]
    row = 2
    for key, label, kind, track in fields:
        ws.cell(row, 1, label).font = LABEL_FONT
        for i, p in enumerate(projections):
            v = p.get(key)
            if isinstance(v, (int, float)):
                cc = ws.cell(row, i + 2, float(v))
                cc.number_format = "0.00\"%\"" if kind == "pct_already" else "#,##0"
                cc.font = BODY_FONT
                # Ghi nhận ô từng năm (vd "fcff_y1", "ebitda_y1", ...) cho công thức.
                if track and refs is not None:
                    refs.record(f"{key}_y{i+1}", ws.title, cc.coordinate)
        row += 1

    if unit:
        ws.cell(row + 1, 1, f"(Đơn vị: {unit}, trừ %)").font = MUTED_FONT
    ws.column_dimensions["A"].width = 32
    for i in range(len(projections)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16
    ws.freeze_panes = "B2"

    # Assumptions block.
    asum = projection.get("assumptions") or {}
    if asum:
        ws.cell(row + 3, 1, "GIẢ ĐỊNH DỰ PHÓNG").font = SECTION_FONT
        rr = row + 4
        for k, v in asum.items():
            ws.cell(rr, 1, str(k)).font = LABEL_FONT
            display = ", ".join(map(str, v)) if isinstance(v, list) else str(v) if v is not None else "—"
            ws.cell(rr, 2, display).font = BODY_FONT
            ws.cell(rr, 2).alignment = Alignment(wrap_text=True, vertical="top")
            rr += 1

    # Summary block.
    summary = projection.get("summary_5y") or {}
    if summary:
        ws.cell(rr + 1, 1, "TÓM TẮT 5 NĂM").font = SECTION_FONT
        rr += 2
        for k, label in [
            ("revenue_cagr_pct", "CAGR doanh thu (%)"),
            ("ebitda_cagr_pct", "CAGR EBITDA (%)"),
            ("fcff_cumulative", "FCFF cộng dồn"),
            ("comments", "Nhận xét"),
        ]:
            v = summary.get(k)
            ws.cell(rr, 1, label).font = LABEL_FONT
            if isinstance(v, (int, float)):
                cc = ws.cell(rr, 2, float(v))
                cc.number_format = "0.00\"%\"" if "pct" in k else "#,##0"
                cc.font = BODY_FONT
            elif v:
                ws.cell(rr, 2, str(v)).font = BODY_FONT
                ws.cell(rr, 2).alignment = Alignment(wrap_text=True, vertical="top")
            rr += 1


def _sheet_valuation(wb, valuation, unit, refs=None):
    """Sheet Định giá với LIVE FORMULAS — đây là 1 financial model thật.

    Layout:
      - INPUTS block (ô vàng): WACC, terminal_growth, total_debt, multiples.
        Sửa giá trị → toàn bộ DCF + Multiples re-compute tự động.
      - DCF block: PV(FCFF_t) = FCFF / (1+WACC)^t  ← formula, FCFF tham chiếu Dự phóng.
      - Multiples block: Equity = EBITDA × mult − Debt  ← formula.
      - Method values: tham chiếu các ô equity ở trên.
      - Fair value summary: MIN / AVG / MAX của method values.

    Khi extractor/projector không có đủ data → fallback giá trị tĩnh từ valuator.py.
    """
    ws = wb.create_sheet("Định giá")
    summary = valuation.get("summary") or {}
    dcf = valuation.get("dcf") or {}
    multiples = valuation.get("multiples") or {}
    assumptions = valuation.get("assumptions") or {}

    formula_font = Font(name="Consolas", size=9, color=_hex(COLORS["primary"]))
    formula_text_font = Font(name="Consolas", size=9, italic=True, color=_hex(COLORS["text_muted"]))

    # ============== INPUTS BLOCK ==============
    # Các ô vàng = edit-được. Sửa → DCF/Multiples tự update.
    _write_section_title(ws, 1, 1, "🟡 INPUTS — Sửa các ô VÀNG để rerun model", span=4)
    ws.cell(2, 1, "Mọi số trong DCF/Multiples bên dưới là FORMULA tham chiếu các ô này. "
                  "Bấm vào ô bất kỳ để xem công thức ở thanh công thức Excel.").font = MUTED_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    inputs_start = 4
    # (label, value, ref_name, format_text, source_note)
    wacc_pct = dcf.get("wacc_pct")  # đã ở dạng % (vd 14.0 = 14%)
    g_pct = dcf.get("terminal_growth_pct")
    debt = dcf.get("debt_subtracted")
    ev_ebitda_mult = (multiples.get("ev_ebitda") or {}).get("multiple")
    pe_mult = (multiples.get("pe") or {}).get("multiple")
    pb_mult = (multiples.get("pb") or {}).get("multiple")

    inputs = [
        ("WACC",                 wacc_pct,       "in_wacc",       "0.00%",       "AI đề xuất, sửa được"),
        ("Terminal growth (g)",  g_pct,          "in_g",          "0.00%",       "AI đề xuất, thường 3-4%"),
        ("Total Debt (trừ ra)",  debt,           "in_debt",       "#,##0",       "Mặc định = total_liabilities BCĐKT"),
        ("EV/EBITDA multiple",   ev_ebitda_mult, "in_ev_ebitda",  "0.00\"x\"",   "AI đề xuất theo ngành"),
        ("P/E multiple",         pe_mult,        "in_pe",         "0.00\"x\"",   "AI đề xuất theo ngành"),
        ("P/B multiple",         pb_mult,        "in_pb",         "0.00\"x\"",   "AI đề xuất theo ngành"),
        ("Minority discount %",  summary.get("minority_discount_pct"), "in_md", "0.00%", "Áp dụng cho SME"),
    ]
    headers = ["Tham số (input)", "Giá trị", "Ghi chú", "Source"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(inputs_start, col, h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    in_refs: dict[str, str] = {}  # local input refs (cùng sheet, viết "B5" thẳng)
    row = inputs_start + 1
    for label, val, key, fmt, note in inputs:
        ws.cell(row, 1, label).font = LABEL_FONT
        c = ws.cell(row, 2)
        if isinstance(val, (int, float)):
            # WACC/g/md là % — input dưới dạng fraction (0.14 thay vì 14) để công thức gọn.
            store_val = val / 100.0 if "%" in fmt else float(val)
            c.value = store_val
            c.number_format = fmt
        c.font = INPUT_FONT
        c.fill = INPUT_FILL  # vàng = ô edit-được
        in_refs[key] = c.coordinate
        ws.cell(row, 3, note).font = MUTED_FONT
        ws.cell(row, 4, "valuator.py · _claude_assumptions").font = formula_text_font
        row += 1

    # Override total_debt input nếu BCĐKT có total_liabilities — biến input thành formula thay vì literal.
    if refs and refs.has("total_liabilities"):
        debt_cell = ws.cell(inputs_start + 1 + 2, 2)  # 3rd input row (Total Debt)
        debt_cell.value = f"={refs.get('total_liabilities')}"
        debt_cell.number_format = "#,##0"
        ws.cell(inputs_start + 1 + 2, 3, "= 'BCĐKT'!total_liabilities (sửa số ở BCĐKT hoặc gõ đè ô này)").font = MUTED_FONT

    row += 1

    # ============== DCF BLOCK ==============
    _write_section_title(ws, row, 1, "DCF (FCFF) — LIVE FORMULAS", span=4)
    row += 1
    ws.cell(row, 1, "📁 Công thức Python: agents/valuator.py · _dcf_valuation (line 157). "
                    "Excel formulas dưới đây EQUIVALENT, có thể tự sửa trong Excel.").font = MUTED_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # PV breakdown table — LIVE formulas tham chiếu Dự phóng FCFF + INPUT WACC.
    pv_breakdown = dcf.get("pv_breakdown") or []
    n_years = len(pv_breakdown) if pv_breakdown else 5
    pv_cells: list[str] = []  # sẽ dùng cho SUM ở dưới

    headers = ["Năm (t)", "FCFF (từ Dự phóng)", "Discount factor", "PV", "Công thức (audit)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h); c.font = LABEL_FONT; c.fill = TOTAL_FILL
    row += 1
    for t in range(1, n_years + 1):
        ws.cell(row, 1, t).font = BODY_FONT
        fcff_ref = refs.get(f"fcff_y{t}") if refs else None

        if fcff_ref:
            # FCFF tham chiếu Dự phóng.
            f_cell = ws.cell(row, 2, f"={fcff_ref}")
            f_cell.number_format = "#,##0"
            f_cell.font = BODY_FONT
            df_cell = ws.cell(row, 3, f"=1/(1+{in_refs['in_wacc']})^{t}")
            df_cell.number_format = "0.0000"
            df_cell.font = BODY_FONT
            pv_cell = ws.cell(row, 4,
                f"={f_cell.coordinate}/(1+{in_refs['in_wacc']})^{t}")
            pv_cell.number_format = "#,##0"
            pv_cell.font = BODY_FONT
            pv_cells.append(pv_cell.coordinate)
            explain = f"= 'Dự phóng 5Y'!FCFF_Y{t} ÷ (1+WACC)^{t}"
        elif pv_breakdown and t <= len(pv_breakdown):
            # Fallback static values nếu không có cell ref.
            it = pv_breakdown[t - 1]
            for col, v, fmt in [(2, it.get("fcff"), "#,##0"),
                                 (3, it.get("discount_factor"), "0.0000"),
                                 (4, it.get("pv"), "#,##0")]:
                if isinstance(v, (int, float)):
                    cc = ws.cell(row, col, float(v)); cc.number_format = fmt; cc.font = BODY_FONT
            if isinstance(it.get("pv"), (int, float)):
                pv_cells.append(_coord(4, row))
            explain = "(static — không có cell ref)"
        else:
            explain = "(không đủ dữ liệu)"

        ec = ws.cell(row, 5, explain)
        ec.font = formula_font
        ec.alignment = Alignment(wrap_text=True, vertical="center")
        row += 1

    # Σ PV(FCFF) row — SUM live formula.
    ws.cell(row, 1, "Σ PV(FCFF) — 5Y").font = LABEL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    sum_pv_cell = ws.cell(row, 4)
    if pv_cells:
        sum_pv_cell.value = f"=SUM({pv_cells[0]}:{pv_cells[-1]})"
    elif isinstance(dcf.get("pv_explicit_fcff"), (int, float)):
        sum_pv_cell.value = float(dcf["pv_explicit_fcff"])
    sum_pv_cell.number_format = "#,##0"
    sum_pv_cell.font = LABEL_FONT
    sum_pv_cell.fill = TOTAL_FILL
    sum_pv_ref = sum_pv_cell.coordinate
    ws.cell(row, 5, "= SUM(PV năm 1..5)").font = formula_font
    row += 2

    # Terminal Value + Equity Value — LIVE formulas.
    last_fcff_ref = refs.get(f"fcff_y{n_years}") if refs else None
    tv_cell_coord = None
    pv_tv_cell_coord = None
    ev_cell_coord = None
    equity_cell_coord = None

    final_rows = []
    if last_fcff_ref:
        final_rows.append(("Terminal Value (TV)",
                           f"={last_fcff_ref}*(1+{in_refs['in_g']})/({in_refs['in_wacc']}-{in_refs['in_g']})",
                           "#,##0",
                           f"= FCFF_Y{n_years} × (1+g) / (WACC − g)  [Gordon Growth]"))
        final_rows.append(("PV(Terminal Value)",
                           "TV_REF/(1+WACC)^N",  # placeholder, replaced after we know TV cell
                           "#,##0",
                           f"= TV ÷ (1+WACC)^{n_years}"))
        final_rows.append(("Enterprise Value (EV)",
                           "SUM_PV+PV_TV",
                           "#,##0",
                           "= Σ PV(FCFF) + PV(TV)"))
        final_rows.append(("Equity Value (DCF)  ⭐",
                           "EV-DEBT",
                           "#,##0",
                           "= EV − Total_Debt"))
    else:
        # Fallback static.
        for label, key, fmt in [
            ("Terminal Value (TV)", "terminal_value", "#,##0"),
            ("PV(Terminal Value)", "pv_terminal", "#,##0"),
            ("Enterprise Value (EV)", "enterprise_value", "#,##0"),
            ("Equity Value (DCF)  ⭐", "equity_value", "#,##0"),
        ]:
            v = dcf.get(key)
            final_rows.append((label, float(v) if isinstance(v, (int, float)) else None,
                               fmt, "(static, không có cell ref)"))

    # Render.
    for i, (label, formula_or_value, fmt, note) in enumerate(final_rows):
        ws.cell(row, 1, label).font = LABEL_FONT
        c = ws.cell(row, 2)

        if last_fcff_ref:
            # Dynamic formula building.
            if i == 0:  # TV
                c.value = formula_or_value
                tv_cell_coord = c.coordinate
            elif i == 1:  # PV(TV)
                c.value = f"={tv_cell_coord}/(1+{in_refs['in_wacc']})^{n_years}"
                pv_tv_cell_coord = c.coordinate
            elif i == 2:  # EV
                c.value = f"={sum_pv_ref}+{pv_tv_cell_coord}"
                ev_cell_coord = c.coordinate
            elif i == 3:  # Equity
                c.value = f"={ev_cell_coord}-{in_refs['in_debt']}"
                equity_cell_coord = c.coordinate
        else:
            if isinstance(formula_or_value, (int, float)):
                c.value = formula_or_value

        c.number_format = fmt
        c.font = LABEL_FONT if i == 3 else BODY_FONT
        if i == 3:
            c.fill = TOTAL_FILL
        ws.cell(row, 3, note).font = formula_font
        row += 1

    if dcf.get("note"):
        ws.cell(row, 1, "Ghi chú").font = LABEL_FONT
        c = ws.cell(row, 2, dcf["note"])
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = MUTED_FONT
        row += 1

    row += 2

    # ============== MULTIPLES BLOCK ==============
    _write_section_title(ws, row, 1, "MULTIPLES — LIVE FORMULAS", span=4)
    row += 1
    ws.cell(row, 1, "📁 agents/valuator.py · _multiples_valuation (line 206). Multiples ở INPUTS block.").font = MUTED_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    headers = ["Phương pháp", "Multiple", "Input cơ sở", "Equity Value", "Công thức"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h); c.font = LABEL_FONT; c.fill = TOTAL_FILL
    row += 1

    # Mỗi multiple: tham chiếu input cell + (FCFF Y1 EBITDA / KQKD net_income / BCĐKT total_equity).
    ev_ebitda_input_ref = refs.get("ebitda_y1") if refs else None
    pe_input_ref = refs.get("net_profit_after_tax") if refs else None
    pb_input_ref = refs.get("total_equity") if refs else None

    multiple_specs = [
        ("EV/EBITDA", in_refs["in_ev_ebitda"], ev_ebitda_input_ref,
         (multiples.get("ev_ebitda") or {}).get("ebitda_input"),
         (multiples.get("ev_ebitda") or {}).get("equity_value"),
         lambda mult, inp: f"={mult}*{inp}-{in_refs['in_debt']}",  # EV/EBITDA: equity = EBITDA*x - debt
         "= EBITDA × multiple − Total_Debt"),
        ("P/E", in_refs["in_pe"], pe_input_ref,
         (multiples.get("pe") or {}).get("net_income_input"),
         (multiples.get("pe") or {}).get("equity_value"),
         lambda mult, inp: f"={mult}*{inp}",  # P/E: equity = NI*x
         "= Net_Income × multiple"),
        ("P/B", in_refs["in_pb"], pb_input_ref,
         (multiples.get("pb") or {}).get("book_value_input"),
         (multiples.get("pb") or {}).get("equity_value"),
         lambda mult, inp: f"={mult}*{inp}",  # P/B: equity = BV*x
         "= Book_Value × multiple"),
    ]
    multiple_equity_cells: dict[str, str] = {}
    for label, mult_ref, input_ref, input_val, fallback_eq, formula_fn, formula_text in multiple_specs:
        ws.cell(row, 1, label).font = BODY_FONT
        # Multiple — formula tham chiếu input cell.
        ws.cell(row, 2, f"={mult_ref}").number_format = "0.00\"x\""
        # Input — formula tham chiếu KQKD/BCĐKT/Dự phóng nếu có; fallback static.
        in_cell = ws.cell(row, 3)
        if input_ref:
            in_cell.value = f"={input_ref}"
        elif isinstance(input_val, (int, float)):
            in_cell.value = float(input_val)
        in_cell.number_format = "#,##0"
        in_cell.font = BODY_FONT
        # Equity = formula(multiple, input).
        eq_cell = ws.cell(row, 4)
        if input_ref or isinstance(input_val, (int, float)):
            eq_cell.value = formula_fn(mult_ref, in_cell.coordinate)
        elif isinstance(fallback_eq, (int, float)):
            eq_cell.value = float(fallback_eq)
        eq_cell.number_format = "#,##0"
        eq_cell.font = BODY_FONT
        multiple_equity_cells[label] = eq_cell.coordinate
        ws.cell(row, 5, formula_text).font = formula_font
        row += 1

    row += 2

    # ============== METHOD VALUES + FAIR VALUE SUMMARY ==============
    _write_section_title(ws, row, 1, "GIÁ TRỊ THEO PHƯƠNG PHÁP (live)", span=2)
    row += 2
    headers = ["Phương pháp", "Equity Value (live)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h); c.font = LABEL_FONT; c.fill = TOTAL_FILL
    row += 1
    method_cells: list[str] = []
    method_specs = []
    if equity_cell_coord:
        method_specs.append(("DCF (FCFF)", f"={equity_cell_coord}"))
    elif isinstance(dcf.get("equity_value"), (int, float)):
        method_specs.append(("DCF (FCFF)", float(dcf["equity_value"])))
    for label_short in ["EV/EBITDA", "P/E", "P/B"]:
        if label_short in multiple_equity_cells:
            method_specs.append((label_short, f"={multiple_equity_cells[label_short]}"))
    for label, val in method_specs:
        ws.cell(row, 1, label).font = BODY_FONT
        c = ws.cell(row, 2, val)
        c.number_format = "#,##0"
        c.font = BODY_FONT
        method_cells.append(c.coordinate)
        row += 1

    if method_cells:
        row += 1
        _write_section_title(ws, row, 1, "FAIR VALUE SUMMARY (live)", span=3)
        row += 2
        first = method_cells[0]; last = method_cells[-1]
        rng = f"{first}:{last}" if len(method_cells) > 1 else first
        fair_specs = [
            ("Fair value low",  f"=MIN({rng})",     "= MIN(các method values)"),
            ("Fair value mid",  f"=AVERAGE({rng})", "= AVERAGE(các method values)"),
            ("Fair value high", f"=MAX({rng})",     "= MAX(các method values)"),
        ]
        fv_cells: dict[str, str] = {}
        for label, formula, formula_text in fair_specs:
            ws.cell(row, 1, label).font = LABEL_FONT
            c = ws.cell(row, 2, formula)
            c.number_format = "#,##0"
            c.font = LABEL_FONT
            fv_cells[label] = c.coordinate
            ws.cell(row, 3, formula_text).font = formula_font
            row += 1
        # Sau chiết khấu thiểu số.
        if "Fair value mid" in fv_cells:
            ws.cell(row, 1, "Sau chiết khấu thiểu số").font = LABEL_FONT
            c = ws.cell(row, 2,
                f"={fv_cells['Fair value mid']}*(1-{in_refs['in_md']})")
            c.number_format = "#,##0"
            c.font = LABEL_FONT
            c.fill = TOTAL_FILL
            ws.cell(row, 3, f"= Fair_value_mid × (1 − minority_discount%)").font = formula_font
            row += 1

    # Assumptions block (text — Claude rationale, không tính).
    if assumptions:
        row += 2
        _write_section_title(ws, row, 1, "RATIONALE GIẢ ĐỊNH (Claude)", span=2)
        row += 2
        for k, v in assumptions.items():
            ws.cell(row, 1, str(k)).font = LABEL_FONT
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cc = ws.cell(row, 2, float(v))
                cc.number_format = "0.00\"%\"" if "pct" in str(k) else "0.00"
                cc.font = BODY_FONT
            elif isinstance(v, dict):
                cc = ws.cell(row, 2, ", ".join(f"{kk}={vv}" for kk, vv in v.items()))
                cc.alignment = Alignment(wrap_text=True, vertical="top")
                cc.font = BODY_FONT
            elif isinstance(v, list):
                cc = ws.cell(row, 2, "; ".join(map(_summarize_item, v)))
                cc.alignment = Alignment(wrap_text=True, vertical="top")
                cc.font = BODY_FONT
            elif v is not None:
                cc = ws.cell(row, 2, str(v))
                cc.alignment = Alignment(wrap_text=True, vertical="top")
                cc.font = BODY_FONT
            row += 1

    if unit:
        ws.cell(row + 2, 1, f"(Đơn vị giá trị: {unit})").font = MUTED_FONT
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 50


def _sheet_sensitivity(wb, valuation):
    sens = valuation.get("sensitivity") or {}
    matrix = sens.get("matrix") or []
    growth_axis = sens.get("growth_axis_pct") or []
    if not matrix or not growth_axis:
        return

    ws = wb.create_sheet("Sensitivity")
    ws.cell(1, 1, "Sensitivity matrix — Equity Value theo WACC × Terminal growth").font = SECTION_FONT
    ws.cell(2, 1, f"Base case: WACC={sens.get('base_wacc_pct')}%, g={sens.get('base_growth_pct')}%").font = MUTED_FONT
    base_w = sens.get("base_wacc_pct")
    base_g = sens.get("base_growth_pct")

    # Header row: g axis.
    ws.cell(4, 1, "WACC \\ g").font = HEADER_FONT
    ws.cell(4, 1).fill = HEADER_FILL
    for col, g in enumerate(growth_axis, 2):
        c = ws.cell(4, col, f"g={g}%")
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # Body rows.
    for r_idx, mrow in enumerate(matrix):
        if not isinstance(mrow, dict):
            continue
        wacc = mrow.get("wacc_pct")
        ws_row = 5 + r_idx
        wacc_cell = ws.cell(ws_row, 1, f"WACC={wacc}%" if wacc is not None else "—")
        wacc_cell.font = LABEL_FONT
        wacc_cell.fill = HEADER_FILL
        wacc_cell.font = HEADER_FONT
        values = mrow.get("values") or []
        for c_idx, val in enumerate(values):
            if not isinstance(val, dict):
                continue
            ev = val.get("equity_value")
            g_val = val.get("terminal_growth_pct")
            cell = ws.cell(ws_row, 2 + c_idx)
            if isinstance(ev, (int, float)):
                cell.value = float(ev)
                cell.number_format = "#,##0"
                cell.font = BODY_FONT
            else:
                cell.value = "—"
                cell.font = MUTED_FONT
            # Highlight base case.
            if (isinstance(wacc, (int, float)) and isinstance(base_w, (int, float)) and abs(wacc - base_w) < 0.01
                    and isinstance(g_val, (int, float)) and isinstance(base_g, (int, float)) and abs(g_val - base_g) < 0.01):
                cell.fill = BASE_HIGHLIGHT_FILL
                cell.font = LABEL_FONT

    ws.column_dimensions["A"].width = 16
    for i in range(len(growth_axis)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18
    ws.freeze_panes = "B5"


def _sheet_industry(wb, industry):
    ws = wb.create_sheet("Phân tích ngành")
    rows = [
        ("Ngành xác định", industry.get("industry_name")),
        ("Cơ sở phân loại", industry.get("industry_classification_basis")),
        ("Tổng quan ngành", industry.get("industry_overview")),
        ("Triển vọng 3 năm", industry.get("industry_outlook_3y")),
        ("CAGR ngành 5Y (%)", industry.get("industry_cagr_5y_pct")),
    ]
    ms = industry.get("market_size") or {}
    rows += [
        ("TAM (tỷ VND)", ms.get("tam_vnd_billion")),
        ("SAM (tỷ VND)", ms.get("sam_vnd_billion")),
        ("SOM (tỷ VND)", ms.get("som_vnd_billion")),
        ("Thị phần DN (%)", ms.get("company_market_share_pct")),
        ("Giả định market size", ms.get("assumptions")),
    ]
    _write_section_title(ws, 1, 1, "TỔNG QUAN NGÀNH", span=2)
    _write_kv_block(ws, rows, start_row=3)
    next_row = 3 + len(rows) + 2

    competitors = industry.get("key_competitors") or []
    if competitors:
        _write_section_title(ws, next_row, 1, "ĐỐI THỦ CẠNH TRANH", span=4)
        next_row += 2
        for col, h in enumerate(["Tên", "Doanh thu ước (tỷ)", "Thị phần (%)", "Ghi chú"], 1):
            c = ws.cell(next_row, col, h); c.font = LABEL_FONT
            c.fill = TOTAL_FILL
        next_row += 1
        for cp in competitors:
            ws.cell(next_row, 1, cp.get("name") or "—").font = BODY_FONT
            v = cp.get("estimated_revenue_vnd_billion")
            if isinstance(v, (int, float)):
                ws.cell(next_row, 2, float(v)).number_format = "#,##0.0"
            v = cp.get("market_share_pct")
            if isinstance(v, (int, float)):
                ws.cell(next_row, 3, float(v)).number_format = "0.0\"%\""
            cc = ws.cell(next_row, 4, cp.get("note") or "")
            cc.alignment = Alignment(wrap_text=True, vertical="top")
            cc.font = BODY_FONT
            next_row += 1

    next_row += 2
    drivers = industry.get("industry_growth_drivers") or []
    risks = industry.get("industry_risks") or []
    barriers = industry.get("barriers_to_entry") or []
    trends = industry.get("industry_trends") or []
    for title, items in [
        ("DRIVER TĂNG TRƯỞNG", drivers),
        ("XU HƯỚNG", trends),
        ("RỦI RO NGÀNH", risks),
        ("RÀO CẢN GIA NHẬP", barriers),
    ]:
        if not items:
            continue
        _write_section_title(ws, next_row, 1, title, span=2)
        next_row += 1
        for it in items:
            ws.cell(next_row, 1, "•").font = BODY_FONT
            c = ws.cell(next_row, 2, str(it))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = BODY_FONT
            next_row += 1
        next_row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 50


def _sheet_thesis(wb, thesis):
    ws = wb.create_sheet("Investment Thesis")
    es = thesis.get("executive_summary") or {}
    it = thesis.get("investment_thesis") or {}
    deal = thesis.get("deal_recommendation") or {}

    _write_section_title(ws, 1, 1, "EXECUTIVE SUMMARY", span=2)
    rows = [
        ("Headline", es.get("headline")),
        ("Khuyến nghị", es.get("recommendation")),
        ("Driver chính", "; ".join(es.get("key_drivers") or [])),
    ]
    _write_kv_block(ws, rows, start_row=3)
    next_row = 3 + len(rows) + 2

    points = it.get("thesis_points") or []
    if points:
        _write_section_title(ws, next_row, 1, "LUẬN ĐIỂM ĐẦU TƯ", span=3)
        next_row += 2
        for col, h in enumerate(["#", "Tiêu đề", "Luận điểm + Bằng chứng"], 1):
            c = ws.cell(next_row, col, h); c.font = LABEL_FONT; c.fill = TOTAL_FILL
        next_row += 1
        for i, p in enumerate(points, 1):
            ws.cell(next_row, 1, i).font = BODY_FONT
            ws.cell(next_row, 2, p.get("title") or "").font = LABEL_FONT
            body = p.get("thesis") or ""
            if p.get("evidence"):
                body += f"\n\nBằng chứng: {p['evidence']}"
            c = ws.cell(next_row, 3, body)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = BODY_FONT
            ws.row_dimensions[next_row].height = 80
            next_row += 1
        next_row += 1

    cats = it.get("catalysts") or []
    if cats:
        _write_section_title(ws, next_row, 1, "CATALYSTS", span=3)
        next_row += 2
        for col, h in enumerate(["Loại", "Mô tả", "Horizon"], 1):
            c = ws.cell(next_row, col, h); c.font = LABEL_FONT; c.fill = TOTAL_FILL
        next_row += 1
        for c in cats:
            ws.cell(next_row, 1, (c.get("type") or "").upper()).font = LABEL_FONT
            cc = ws.cell(next_row, 2, c.get("description") or "")
            cc.alignment = Alignment(wrap_text=True, vertical="top"); cc.font = BODY_FONT
            ws.cell(next_row, 3, c.get("horizon") or "").font = BODY_FONT
            ws.cell(next_row, 1).fill = GOOD_FILL
            next_row += 1
        next_row += 1

    risks = it.get("risks") or []
    if risks:
        _write_section_title(ws, next_row, 1, "RỦI RO", span=4)
        next_row += 2
        for col, h in enumerate(["Loại", "Mức độ", "Mô tả", "Mitigation"], 1):
            c = ws.cell(next_row, col, h); c.font = LABEL_FONT; c.fill = TOTAL_FILL
        next_row += 1
        for r in risks:
            sev = (r.get("severity") or "").upper()
            ws.cell(next_row, 1, (r.get("type") or "").upper()).font = LABEL_FONT
            sev_cell = ws.cell(next_row, 2, sev)
            sev_cell.font = LABEL_FONT
            if sev == "HIGH":
                sev_cell.fill = POOR_FILL
            elif sev == "MEDIUM":
                sev_cell.fill = WARN_FILL
            elif sev == "LOW":
                sev_cell.fill = GOOD_FILL
            cc = ws.cell(next_row, 3, r.get("description") or "")
            cc.alignment = Alignment(wrap_text=True, vertical="top"); cc.font = BODY_FONT
            cc = ws.cell(next_row, 4, r.get("mitigation") or "")
            cc.alignment = Alignment(wrap_text=True, vertical="top"); cc.font = BODY_FONT
            next_row += 1
        next_row += 1

    if deal:
        _write_section_title(ws, next_row, 1, "KHUYẾN NGHỊ DEAL", span=2)
        next_row += 2
        deal_rows = [
            ("Mục tiêu", deal.get("primary_objective")),
            ("Khoảng giá trị", deal.get("fair_value_range_text")),
            ("Entry price", deal.get("entry_price_recommendation")),
            ("Cấu trúc deal", deal.get("deal_structure")),
            ("Bước tiếp theo", "\n".join(f"{i}. {s}" for i, s in enumerate(deal.get("next_steps") or [], 1))),
        ]
        _write_kv_block(ws, deal_rows, start_row=next_row)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 64
    ws.column_dimensions["D"].width = 40


# ---------- Helpers ----------
def _write_section_title(ws, row, col, text, span=1):
    c = ws.cell(row, col, text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        for k in range(1, span):
            ws.cell(row, col + k).fill = HEADER_FILL


def _write_kv_block(ws, rows, start_row, value_format=None):
    for i, (k, v) in enumerate(rows):
        ws.cell(start_row + i, 1, k).font = LABEL_FONT
        if v is None or v == "":
            ws.cell(start_row + i, 2, "—").font = MUTED_FONT
        elif isinstance(v, (int, float)) and not isinstance(v, bool):
            cc = ws.cell(start_row + i, 2, float(v))
            cc.number_format = value_format or "#,##0.00"
            cc.font = BODY_FONT
        else:
            cc = ws.cell(start_row + i, 2, str(v))
            cc.alignment = Alignment(wrap_text=True, vertical="top")
            cc.font = BODY_FONT


def _write_financial_table(ws, fields, cur, prev, cur_label, prev_label,
                            refs=None, sheet_title=None):
    """Write a 5-column financial table với LIVE formulas cho cột Δ.

    fields: list of (key, label) hoặc (key, label, ref_name).
            ref_name: nếu set, ghi nhận cell ref vào `refs` để các sheet khác
            tham chiếu công thức.
    """
    headers = ["Khoản mục", cur_label, prev_label, "Δ tuyệt đối", "Δ %"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center" if col > 1 else "left")

    for i, item in enumerate(fields, 2):
        # Hỗ trợ cả tuple 2 và 3 phần tử (backwards compat).
        if len(item) == 2:
            key, label = item
            ref_name = None
        else:
            key, label, ref_name = item[0], item[1], item[2]
        cv = cur.get(key)
        pv = prev.get(key)
        ws.cell(i, 1, label).font = BODY_FONT

        cv_cell = None; pv_cell = None
        if isinstance(cv, (int, float)):
            cv_cell = ws.cell(i, 2, float(cv))
            cv_cell.number_format = "#,##0"
            cv_cell.font = BODY_FONT
        if isinstance(pv, (int, float)):
            pv_cell = ws.cell(i, 3, float(pv))
            pv_cell.number_format = "#,##0"
            pv_cell.font = BODY_FONT

        # Δ tuyệt đối + Δ % là FORMULA — kế toán bấm vào ô sẽ thấy =B-C, =(B-C)/ABS(C).
        if cv_cell is not None and pv_cell is not None:
            d_abs = ws.cell(i, 4, f"={cv_cell.coordinate}-{pv_cell.coordinate}")
            d_abs.number_format = "#,##0"
            d_abs.font = BODY_FONT
            d_pct = ws.cell(i, 5,
                f"=IFERROR(({cv_cell.coordinate}-{pv_cell.coordinate})"
                f"/ABS({pv_cell.coordinate}),\"\")")
            d_pct.number_format = "+0.0%;-0.0%;0.0%"
            d_pct.font = BODY_FONT

        # Track cell ref cho sheet khác tham chiếu (formula bar sẽ hiện cross-sheet).
        if ref_name and refs is not None and sheet_title and cv_cell is not None:
            refs.record(ref_name, sheet_title, cv_cell.coordinate)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12


def _summarize_item(item):
    if isinstance(item, dict):
        # Most common: comparable companies — show name only.
        for key in ("name", "title", "label"):
            if item.get(key):
                return str(item[key])
        return ", ".join(f"{k}={v}" for k, v in item.items() if v is not None)[:80]
    return str(item)


def _flatten_ratios(rated):
    """Convert analyzer's nested {category: {name: {value, rating}}} → flat {name: value}.

    analyzer.py emits ratios under 4 categories (liquidity / leverage / profitability /
    efficiency). This sheet displays them by category but accesses values flatly.
    """
    flat: dict = {}
    if not isinstance(rated, dict):
        return flat
    for cat, items in rated.items():
        if not isinstance(items, dict):
            continue
        for name, payload in items.items():
            if isinstance(payload, dict):
                flat[name] = payload.get("value")
            else:
                flat[name] = payload
    return flat


# ---------- Methodology sheet ----------
# (Calculation, Type, Formula, Source file:function — line gần đúng).
# Sửa file/line ở đây nếu code Python được di chuyển sau này.
_METHODOLOGY_ROWS = [
    ("AGENT 1 — TRÍCH XUẤT BCTC", None, None, None),
    ("Đọc PDF/ảnh BCTC → JSON", "AI",
     "Multimodal Claude Opus 4.7 + extended thinking. Không có phép tính.",
     "agents/extractor.py · _PROMPT (line 22)"),

    ("AGENT 2 — PHÂN TÍCH NGÀNH", None, None, None),
    ("Industry name, TAM/SAM/SOM, đối thủ", "AI",
     "Phán đoán dựa trên kiến thức ngành. Không có công thức.",
     "agents/industry.py · user_prompt"),

    ("AGENT 3 — TỔNG QUAN DN", None, None, None),
    ("Business model, value chain, vị thế", "AI",
     "Suy luận từ BCTC + ngành.",
     "agents/business_profile.py · user_prompt"),

    ("AGENT 4 — TỶ SỐ TÀI CHÍNH", None, None, None),
    ("current_ratio", "PYTHON",
     "= current_assets_total / current_liabilities_total",
     "agents/analyzer.py · _compute_period_ratios (line 125)"),
    ("quick_ratio", "PYTHON",
     "= (current_assets_total − inventory) / current_liabilities_total",
     "agents/analyzer.py · _compute_period_ratios (line 126)"),
    ("cash_ratio", "PYTHON",
     "= (cash_and_equivalents + short_term_investments) / current_liabilities_total",
     "agents/analyzer.py · _compute_period_ratios (line 127)"),
    ("debt_ratio", "PYTHON",
     "= total_liabilities / total_assets",
     "agents/analyzer.py · _compute_period_ratios (line 130)"),
    ("debt_to_equity", "PYTHON",
     "= total_liabilities / total_equity",
     "agents/analyzer.py · _compute_period_ratios (line 131)"),
    ("equity_multiplier", "PYTHON",
     "= total_assets / total_equity",
     "agents/analyzer.py · _compute_period_ratios (line 132)"),
    ("interest_coverage", "PYTHON",
     "= operating_profit / interest_expense  (EBIT/Interest)",
     "agents/analyzer.py · _compute_period_ratios (line 133)"),
    ("debt_to_ebitda", "PYTHON",
     "= total_liabilities / EBITDA   (EBIT proxy nếu BCTC thiếu D&A — sửa biến `ebitda_proxy`)",
     "agents/analyzer.py · _compute_period_ratios (line 113, 134)"),
    ("gross_margin", "PYTHON",
     "= gross_profit / net_revenue",
     "agents/analyzer.py · _compute_period_ratios (line 137)"),
    ("operating_margin", "PYTHON",
     "= operating_profit / net_revenue",
     "agents/analyzer.py · _compute_period_ratios (line 138)"),
    ("ebitda_margin", "PYTHON",
     "= EBITDA / net_revenue   (EBIT proxy nếu thiếu D&A)",
     "agents/analyzer.py · _compute_period_ratios (line 139)"),
    ("net_margin", "PYTHON",
     "= net_profit_after_tax / net_revenue",
     "agents/analyzer.py · _compute_period_ratios (line 140)"),
    ("roa", "PYTHON",
     "= net_profit_after_tax / total_assets   (cuối kỳ, không trung bình)",
     "agents/analyzer.py · _compute_period_ratios (line 141)"),
    ("roe", "PYTHON",
     "= net_profit_after_tax / total_equity   (cuối kỳ)",
     "agents/analyzer.py · _compute_period_ratios (line 142)"),
    ("asset_turnover", "PYTHON",
     "= net_revenue / total_assets",
     "agents/analyzer.py · _compute_period_ratios (line 145)"),
    ("inventory_turnover", "PYTHON",
     "= cogs / inventory",
     "agents/analyzer.py · _compute_period_ratios (line 146)"),
    ("receivables_turnover", "PYTHON",
     "= net_revenue / short_term_receivables",
     "agents/analyzer.py · _compute_period_ratios (line 147)"),
    ("Tăng trưởng YoY (mọi chỉ tiêu)", "PYTHON",
     "= (current − previous) / |previous|",
     "agents/analyzer.py · _yoy (line 188), _compute_growth (line 200)"),
    ("Ngưỡng đánh giá tốt/cảnh báo/kém", "PYTHON",
     "Bảng ngưỡng cho từng tỷ số (cao tốt / thấp tốt). Sửa để đổi rating.",
     "agents/analyzer.py · _THRESHOLDS (line 43)"),

    ("AGENT 5 — DỰ PHÓNG 5 NĂM", None, None, None),
    ("revenue Y_t", "AI có ràng buộc",
     "= revenue_{t-1} × (1 + revenue_growth_pct[t])",
     "agents/projector.py · prompt line 62"),
    ("gross_profit Y_t", "AI có ràng buộc",
     "= revenue × gross_margin_pct[t]",
     "agents/projector.py · prompt line 64"),
    ("EBIT Y_t", "AI có ràng buộc",
     "= gross_profit − operating_expense  (operating_expense = revenue × OPEX%)",
     "agents/projector.py · prompt line 65"),
    ("EBITDA Y_t", "AI có ràng buộc",
     "= EBIT + Depreciation",
     "agents/projector.py · prompt line 92"),
    ("net_income Y_t", "AI có ràng buộc",
     "= (EBIT − interest_expense) × (1 − tax_rate_pct/100)",
     "agents/projector.py · prompt line 87-88"),
    ("FCFF Y_t   ⭐ quan trọng cho DCF", "AI có ràng buộc",
     "= EBIT × (1 − tax%) + D&A − CAPEX − ΔWC",
     "agents/projector.py · prompt line 110 (ràng buộc bắt buộc)"),
    ("Tăng trưởng giảm dần (S-curve)", "AI có ràng buộc",
     "Y1 cao nhất, Y5 thấp nhất, hội tụ về CAGR ngành",
     "agents/projector.py · prompt line 108"),

    ("AGENT 6 — ĐỊNH GIÁ", None, None, None),
    ("WACC, terminal_growth, multiples", "AI",
     "Claude đề xuất giá trị + rationale. Sửa prompt để force range khác (vd 12-18%).",
     "agents/valuator.py · _claude_assumptions (line 68)"),
    ("DCF: PV(FCFF_t)   ⭐", "PYTHON",
     "= FCFF_t / (1 + WACC)^t      với t = 1..5",
     "agents/valuator.py · _dcf_valuation (line 174-178)"),
    ("DCF: Terminal Value", "PYTHON",
     "= FCFF_5 × (1 + g) / (WACC − g)     [Gordon Growth Model]",
     "agents/valuator.py · _dcf_valuation (line 180-182)"),
    ("DCF: PV(Terminal Value)", "PYTHON",
     "= TV / (1 + WACC)^5",
     "agents/valuator.py · _dcf_valuation (line 183)"),
    ("DCF: Enterprise Value", "PYTHON",
     "= Σ PV(FCFF_t) + PV(TV)",
     "agents/valuator.py · _dcf_valuation (line 188)"),
    ("DCF: Equity Value", "PYTHON",
     "= Enterprise_Value − Total_Debt",
     "agents/valuator.py · _dcf_valuation (line 189)"),
    ("Total_Debt định nghĩa", "PYTHON",
     "= total_liabilities (toàn bộ nợ phải trả). Đổi sang chỉ short_term_debt + long_term_debt nếu muốn 'nợ có lãi vay' thuần.",
     "agents/valuator.py · line 28"),
    ("Multiples — EV/EBITDA", "PYTHON",
     "Equity = (EBITDA_Y1 × ev_ebitda_multiple) − Total_Debt",
     "agents/valuator.py · _multiples_valuation (line 208-216)"),
    ("Multiples — P/E", "PYTHON",
     "Equity = Net_Income_current × pe_multiple",
     "agents/valuator.py · _multiples_valuation (line 219-224)"),
    ("Multiples — P/B", "PYTHON",
     "Equity = Total_Equity_book × pb_multiple",
     "agents/valuator.py · _multiples_valuation (line 227-232)"),
    ("Fair Value low/mid/high", "PYTHON",
     "low = min(equity_values), high = max, mid = mean (TRUNG BÌNH CỘNG, không trọng số)",
     "agents/valuator.py · _build_summary (line 286-289)"),
    ("Chiết khấu thiểu số", "PYTHON",
     "fair_value_after_md = fair_value_mid × (1 − minority_discount_pct/100)",
     "agents/valuator.py · _build_summary (line 291-293)"),
    ("Sensitivity matrix (5×5)", "PYTHON",
     "Re-run DCF với WACC ± 1%, ±2% và g ± 1%, ±2%. Step size sửa được.",
     "agents/valuator.py · _sensitivity (line 240-261)"),

    ("AGENT 7 — INVESTMENT THESIS", None, None, None),
    ("Headline, drivers, catalysts, risks", "AI",
     "Claude tổng hợp output A1-A6 thành memo. Không có phép tính.",
     "agents/thesis_writer.py · user_prompt"),

    ("AGENT 8 — RENDER", None, None, None),
    ("PDF (matplotlib) + Excel (openpyxl)", "PYTHON (vẽ)",
     "Không tạo số mới — chỉ vẽ output. Style tokens tách riêng.",
     "agents/renderer.py · agents/excel_writer.py · style: agents/report_style.py"),
]


def _sheet_methodology(wb):
    """First sheet: bảng tra cứu mọi phép tính trong báo cáo."""
    ws = wb.create_sheet("Phương pháp tính")
    ws.cell(1, 1, "BẢNG TRA CỨU PHƯƠNG PHÁP TÍNH").font = Font(
        name="Calibri", size=14, bold=True, color=_hex(COLORS["primary"]))
    ws.cell(2, 1, "Mỗi dòng = 1 phép tính. Cột 'File nguồn' chỉ ra chỗ sửa nếu cần.").font = MUTED_FONT
    ws.cell(3, 1, "PYTHON = công thức cố định, kế toán có thể audit.   "
                  "AI = phán đoán bởi Claude — đọc rationale trong trace.json.").font = MUTED_FONT

    headers = ["Phép tính / Chỉ tiêu", "Loại", "Công thức / Diễn giải", "File nguồn (Sửa tại đây)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(5, col, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")

    row = 6
    formula_font = Font(name="Consolas", size=9, color=_hex(COLORS["primary"]))
    source_font = Font(name="Consolas", size=9, color=_hex(COLORS["accent"]))
    for entry in _METHODOLOGY_ROWS:
        name, kind, formula, source = entry
        if kind is None:
            # Section header.
            c = ws.cell(row, 1, name)
            c.font = SECTION_FONT
            c.fill = SECTION_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            for k in range(2, 5):
                ws.cell(row, k).fill = SECTION_FILL
            row += 1
            continue
        ws.cell(row, 1, name).font = LABEL_FONT
        kind_cell = ws.cell(row, 2, kind)
        kind_cell.font = LABEL_FONT
        if kind == "PYTHON":
            kind_cell.fill = GOOD_FILL
        elif kind.startswith("AI"):
            kind_cell.fill = PatternFill("solid", fgColor=_hex(COLORS.get("primary_band") or "DBEAFE"))
        else:
            kind_cell.fill = TOTAL_FILL
        f_cell = ws.cell(row, 3, formula or "")
        f_cell.font = formula_font
        f_cell.alignment = Alignment(wrap_text=True, vertical="top")
        s_cell = ws.cell(row, 4, source or "")
        s_cell.font = source_font
        s_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 28
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 56
    ws.freeze_panes = "A6"
